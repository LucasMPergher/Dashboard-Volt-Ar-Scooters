"""Pruebas del generador de datos semanales simulados."""

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.config import (
    ANTIGUEDAD_MAXIMA_MESES,
    ANTIGUEDAD_MINIMA_MESES,
    AUTONOMIA_MAXIMA_KM,
    AUTONOMIA_MINIMA_KM,
    CATEGORIAS_NIVEL_FALLOS,
    CATEGORIAS_SUCURSAL,
    MUESTRA_PREDETERMINADA,
    SEMILLA_PREDETERMINADA,
    VARIABLES_ESTADISTICAS,
    VARIABLE_ANTIGUEDAD_BATERIA,
    VARIABLE_AUTONOMIA_REAL,
    VARIABLE_NIVEL_FALLOS,
    VARIABLE_SUCURSAL,
)
from src.simulacion_datos import (
    calcular_frecuencias_esperadas,
    construir_ruta_salida,
    generar_datos,
    guardar_excel,
    validar_control_predeterminado,
)


def test_generacion_predeterminada_tiene_48_filas() -> None:
    """Comprueba la cantidad predeterminada de observaciones."""
    datos = generar_datos()

    assert len(datos) == MUESTRA_PREDETERMINADA


def test_generacion_tiene_exactamente_cuatro_columnas() -> None:
    """Comprueba que no se agreguen columnas extra."""
    datos = generar_datos()

    assert datos.shape[1] == 4


def test_orden_exacto_de_columnas() -> None:
    """Comprueba el orden requerido de las columnas estadísticas."""
    datos = generar_datos()

    assert tuple(datos.columns) == VARIABLES_ESTADISTICAS


def test_reproducibilidad_con_misma_semilla() -> None:
    """Misma cantidad y semilla deben producir el mismo DataFrame."""
    datos_1 = generar_datos(cantidad=48, semilla=42)
    datos_2 = generar_datos(cantidad=48, semilla=42)

    pd.testing.assert_frame_equal(datos_1, datos_2)


def test_resultados_distintos_con_semillas_diferentes() -> None:
    """Semillas diferentes deben producir datos distintos."""
    datos_1 = generar_datos(cantidad=48, semilla=42)
    datos_2 = generar_datos(cantidad=48, semilla=43)

    assert not datos_1.equals(datos_2)


def test_rechaza_menos_de_30_observaciones() -> None:
    """El generador debe rechazar muestras menores al mínimo."""
    with pytest.raises(ValueError, match="entre 30 y 60"):
        generar_datos(cantidad=29)


def test_rechaza_mas_de_60_observaciones() -> None:
    """El generador debe rechazar muestras mayores al máximo."""
    with pytest.raises(ValueError, match="entre 30 y 60"):
        generar_datos(cantidad=61)


def test_no_contiene_valores_nulos() -> None:
    """Los datos generados no deben tener nulos."""
    datos = generar_datos()

    assert not datos.isna().any().any()


def test_sucursales_validas() -> None:
    """La variable Sucursal debe usar solo categorías válidas."""
    datos = generar_datos()

    assert set(datos[VARIABLE_SUCURSAL]).issubset(CATEGORIAS_SUCURSAL)


def test_niveles_de_fallos_validos() -> None:
    """La variable Nivel_Fallos debe usar solo categorías válidas."""
    datos = generar_datos()

    assert set(datos[VARIABLE_NIVEL_FALLOS]).issubset(CATEGORIAS_NIVEL_FALLOS)


def test_antiguedad_es_entera() -> None:
    """La antigüedad de batería debe ser entera."""
    datos = generar_datos()

    assert pd.api.types.is_integer_dtype(datos[VARIABLE_ANTIGUEDAD_BATERIA])


def test_rango_de_antiguedad() -> None:
    """La antigüedad debe estar entre 1 y 48 meses."""
    datos = generar_datos()

    assert datos[VARIABLE_ANTIGUEDAD_BATERIA].between(
        ANTIGUEDAD_MINIMA_MESES,
        ANTIGUEDAD_MAXIMA_MESES,
    ).all()


def test_rango_de_autonomia() -> None:
    """La autonomía debe estar entre 15 y 45 kilómetros."""
    datos = generar_datos()

    assert datos[VARIABLE_AUTONOMIA_REAL].between(
        AUTONOMIA_MINIMA_KM,
        AUTONOMIA_MAXIMA_KM,
    ).all()


def test_variabilidad_de_variables_cuantitativas() -> None:
    """Las variables cuantitativas deben presentar variabilidad."""
    datos = generar_datos()

    assert datos[VARIABLE_ANTIGUEDAD_BATERIA].nunique() > 1
    assert datos[VARIABLE_AUTONOMIA_REAL].nunique() > 1


def test_correlacion_negativa_y_no_perfecta() -> None:
    """La relación entre antigüedad y autonomía debe ser negativa y no perfecta."""
    datos = generar_datos(
        cantidad=MUESTRA_PREDETERMINADA,
        semilla=SEMILLA_PREDETERMINADA,
    )
    correlacion = datos[VARIABLE_ANTIGUEDAD_BATERIA].corr(
        datos[VARIABLE_AUTONOMIA_REAL]
    )

    assert -0.95 <= correlacion <= -0.55
    assert correlacion != -1


def test_balance_entre_sucursales() -> None:
    """La diferencia entre sucursales no debe superar una observación."""
    datos = generar_datos(cantidad=49, semilla=42)
    conteos = datos[VARIABLE_SUCURSAL].value_counts()

    assert conteos.max() - conteos.min() <= 1


def test_frecuencias_esperadas_del_conjunto_predeterminado_son_robustas() -> None:
    """Controla frecuencias esperadas sin implementar Chi-cuadrado definitivo."""
    datos = generar_datos()
    esperadas = validar_control_predeterminado(datos)
    proporcion_mayores_igual_cinco = (esperadas >= 5).to_numpy().mean()

    assert (esperadas >= 1).all().all()
    assert proporcion_mayores_igual_cinco >= 0.80


def test_escritura_correcta_del_excel(tmp_path) -> None:
    """El archivo Excel debe escribirse en la ruta solicitada."""
    datos = generar_datos()
    ruta = tmp_path / "volt_ar_semana_01.xlsx"

    resultado = guardar_excel(datos, ruta)

    assert resultado == ruta
    assert ruta.exists()


def test_lectura_del_excel_conserva_dimensiones_y_columnas(tmp_path) -> None:
    """La lectura del Excel debe conservar dimensiones y columnas."""
    datos = generar_datos()
    ruta = guardar_excel(datos, tmp_path / "volt_ar_semana_01.xlsx")

    leidos = pd.read_excel(ruta, sheet_name="datos")

    assert leidos.shape == datos.shape
    assert tuple(leidos.columns) == VARIABLES_ESTADISTICAS


def test_nombre_de_hoja_es_datos(tmp_path) -> None:
    """El Excel debe contener una hoja llamada datos."""
    datos = generar_datos()
    ruta = guardar_excel(datos, tmp_path / "volt_ar_semana_01.xlsx")
    libro = load_workbook(ruta, read_only=True)

    assert libro.sheetnames == ["datos"]


def test_rechaza_semana_no_positiva() -> None:
    """La semana usada en el nombre del archivo debe ser positiva."""
    with pytest.raises(ValueError, match="entero positivo"):
        construir_ruta_salida(semana=0)


def test_conjunto_predeterminado_tiene_tres_niveles_de_fallos() -> None:
    """La configuración predeterminada debe incluir Bajo, Medio y Alto."""
    datos = generar_datos()
    frecuencias = calcular_frecuencias_esperadas(datos)

    assert set(datos[VARIABLE_NIVEL_FALLOS]) == set(CATEGORIAS_NIVEL_FALLOS)
    assert frecuencias.shape == (2, 3)
